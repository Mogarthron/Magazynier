from flask import render_template, request, redirect, url_for, send_file
from ..magazyny_pianki import magazyny_pianki


magazyny_pianek = {
    "A": {"nazwa": "MAGAZYN 1 - ANTRESOLA 1"},
    "B": {"nazwa": "MAGAZYN 1 - ANTRESOLA 2"},
    # "MAGAZYN 2": {"skr":"C", },
    # "MAGAZYN 10": {"skr":"D", },
    # "MAGAZYN 11": {"skr":"E", }
            }


import fdb
import pandas as pd
import numpy as np

from Modele_db import session
from Modele_db.modele_db import ZAM_PIANKI, KOMPLETY_PIANEK

import json

with open("linki.json", "r") as c:
    param = json.load(c)["firebird"]


def pobierz_dane_do_kompletacji_stelarzy(nr_paczki):

    con = fdb.connect(dsn=param["dns"], user=param["user"], password=param["password"], sql_dialect=param["sql_dialect"], fb_library_name=param["fb_library_name"])

    zam_spec = pd.read_sql(f"""SELECT 
                                NR_KOMISJI, 
                                NR_REFERENCYJNY AS ARTYKUL_OPIS, 
                                ARTYKUL_KOD, 
                                ILOSC_ZAM, 
                                --LP, 
                                PP, 
                                ZAMOWIENIA_SPEC_ID, 
                                ZAMOWIENIA_ID 
                                FROM ZAMOWIENIA_SPEC 
                                WHERE PP LIKE '%{nr_paczki}%' AND 
                                ARTYKUL_KOD LIKE '10.%' AND 
                                ARTYKUL_KOD NOT LIKE '10.800%'""", con=con)


    con.close()

    glowna_paczka = zam_spec[(~zam_spec.PP.str.contains("MAG"))&(~zam_spec.ARTYKUL_OPIS.str.contains(" PD |POKROWIEC|PODUSZKA|STOLIK|ARTYKU|MATERIA|ŁĄCZNIK|SAMPLER|0/"))].copy()


    glowna_paczka["RODZINA"] = glowna_paczka.ARTYKUL_KOD.str[2:7].map({
                                        ".135.": "AMA",
                                        ".009.": "AVA",
                                        ".131.": "CAL",
                                        ".117.": "DIV",
                                        ".127.": "DIV",
                                        ".022.": "DUO",
                                        ".023.": "DUO",
                                        ".086.": "ELI",
                                        ".132.": "EXT",
                                        ".105.": "HOR",
                                        ".104.": "HOR",
                                        ".128.": "HOR",
                                        ".129.": "HUD",
                                        ".139.": "MAX",
                                        ".093.": "MYS",
                                        ".138.": "OXY",
                                        ".116.": "ONY",
                                        ".133.": "REV",
                                        ".115.": "RIT",
                                        ".077.": "SAM",
                                        ".118.": "SPE",
                                        ".125.": "STO",
                                        ".137.": "CUP",
                                        ".130.": "WIL",
                                        #  ".040.": "OVL",
                                        #  ".121.": "LEN",
                                        #  ".126.": "GRE",
                                        #  ".136.": "KEL",
                                        # #  ".135.": "UNO",
                                        # #  ".134.": "UNO",
                                        #".140.": "COC",
                                        ".141.": "GOY"})


    def unikalne_komisje(kom):
        return kom.drop_duplicates().count()

    podsymowanie_paczki = glowna_paczka.groupby("RODZINA").agg({"NR_KOMISJI":unikalne_komisje ,"ARTYKUL_OPIS": "count"})


    magazyny = pd.DataFrame(data=[
        ["AVA",	"MAG 11",	"MAG 01",	3.5],
        ["CAL",	"MAG 16",	"MAG 02",	3.2],
        ["DIV",	"MAG 17",	"MAG 10",	3.8],
        ["DUO",	"MAG 16",	"MAG 02",	3.2],
        ["ELI",	"MAG 11",	"MAG 01",	3.6],
        ["GRE",	"MAG 01",	"MAG 02",	0],
        ["HOR",	"MAG 11",	"MAG 01",	4.25],
        ["HUD",	"MAG 11",	"MAG 02",	3.2],
        ["LEN",	"MAG 17",	"MAG 01",	0],
        ["ONY",	"MAG 16",	"MAG 02",	3.2],
        ["OVA",	"MAG 01",	"MAG 02",	0],
        ["RIT",	"MAG 17",	"MAG 10",	5],
        ["SAM",	"MAG 16",	"MAG 02",	3.2],
        ["SPE",	"MAG 16",	"MAG 02",	4],
        ["STO",	"MAG 15",	"MAG 01",	4.25],
        ["WIL",	"MAG 16",	"MAG 02",	3.9],
        ["EXT",	"MAG 17",	"MAG 02",	3.2],
        ["REV",	"MAG 10",	"MAG 02",	4.4],
        ["UNO",	"MAG 01",	"MAG 01",	0],
        ["KEL",	"MAG 01",	"MAG 01",	0],
        ["AMA",	"MAG 15",	"MAG 01",	4.4],
        ["TOB",	"MAG 17",	"MAG 10",	3.8],
        ["CUP",	"MAG 15",	"MAG 01",	4.25],
        ["COC",	"MAG 01",	"MAG 01",	0],
        ["MAX",	"MAG 11",	"MAG 02",	4],
        ["OXY",	"MAG 10",	"MAG 02",	4],
        ["GOY",	"MAG 17",	"MAG 10",	4.4]
    ], columns=["RODZINA","MAGAZYN_POBRANIA_STELAZY","MAGAZYN_PIANKI","CZAS_KOMPLETACJI"])


    df = podsymowanie_paczki.merge(magazyny, how="left", on="RODZINA")


    df["CZAS_NA_PROCES"] = df.ARTYKUL_OPIS*df.CZAS_KOMPLETACJI
    df_gb = df.groupby("MAGAZYN_PIANKI")["ARTYKUL_OPIS"].sum().reset_index()

    df_gb["SREDNI_CZAS_KOMPL"] = np.array([1.4,0.9,1.1])
    df_gb["PRZYBLIZONY_CZAS_KOMPL"] = df_gb.ARTYKUL_OPIS * df_gb.SREDNI_CZAS_KOMPL


    return podsymowanie_paczki, df, df_gb, glowna_paczka

@magazyny_pianki.route("/", methods=["POST", "GET"])
def index():

    if request.method == "POST":
        print(request.form["nr_paczki"])
        return redirect(url_for("magazyny_pianki.kompletacje_stelarzy", nr_paczki=request.form["nr_paczki"].replace("/", "_")))


    return render_template("magazyny_pianki.html", title="MAGAZYNY PIANEK", magazyny_pianek=magazyny_pianek)

@magazyny_pianki.route("/kompletacja_stelarzy/<nr_paczki>", methods=["POST", "GET"])
def kompletacje_stelarzy(nr_paczki):
    
    nr_paczki = nr_paczki.replace("_", "/")

    
    _, df, df_gb, _ = pobierz_dane_do_kompletacji_stelarzy(nr_paczki)

    return render_template("kompletacja_stelarzy.html", title="KOMPLETACJA STELARZY", tabelka_z_kompletacja=df, mag_podsum=df_gb)

@magazyny_pianki.route("/pobierz_raport/<nr_paczki>")
def pobierz_raport(nr_paczki):

    nr_paczki = nr_paczki.replace("_", "/")
    

    _, df, _, glowna_paczka = pobierz_dane_do_kompletacji_stelarzy(nr_paczki)

    import openpyxl
    from openpyxl.utils.dataframe import dataframe_to_rows


    wb = openpyxl.Workbook()
    wb.create_sheet(f"PP0{nr_paczki.split('/')[1]}")


    for r in dataframe_to_rows(df, index=False, header=True):
        wb[f"PP0{nr_paczki.split('/')[1]}"].append(r)


    for rod in df.RODZINA.to_list():
        wb.create_sheet(rod)

        for r in dataframe_to_rows(glowna_paczka[glowna_paczka.RODZINA == rod], index=False, header=True):
            wb[rod].append(r)

    wb.save(f"{nr_paczki.replace('/', '_')}.xlsx")

  
    return send_file(f"../{nr_paczki.replace('/', '_')}.xlsx")

@magazyny_pianki.route("/magazyn/<ozn_mag>")
def magazyn(ozn_mag):

    pass