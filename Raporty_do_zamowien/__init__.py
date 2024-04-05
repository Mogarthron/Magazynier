import pandas as pd
import numpy as np
from Modele_db.modele_db import *
from Analiza_pianek import funkcje_analizy_pianek as ap, owaty as ow
from Modele_pianek import tab
from Analiza_pianek.instrukcje_zamawiana import instrukcja_zamawiania_pianpol as ar

ard = {a.MODEL: a for a in ar}
komplety_pianek = ap.komplety_pianek
zam_pianki = ap.zam_pianki
_owaty = ow._owaty

import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side
from datetime import datetime as dt
import json

with open("linki.json", "r", encoding="UTF8") as f:
  rozkroj_pianek = json.load(f)["rozkroj_pianek"]


#zlecenia na kopletacje pianek i owat
def zlecenia_produkcyjne_pianki_owaty(model, nr_kompletacji, nr_partii):
  """
  nr_partii -> numer tygodnia / np.: 07/1
  """
  zp = zam_pianki[(zam_pianki.OPIS.str.contains(model))&(zam_pianki.NR_KOMPLETACJI == nr_kompletacji)][["OPIS", "ILE_ZAMOWIONE"]]

  for p in range(zp.shape[0]):

    ow = _owaty[_owaty.OPIS == zp.iloc[p].OPIS][["TYP_OWATY", "ZUZYCIE", "RODZAJ_CIECIA", "NAZWA_UKL"]].reset_index()
    ow["ZUZYCIE"] = ow.ZUZYCIE*zp.iloc[p].ILE_ZAMOWIONE*1.1
    ow["KATER_UKL"] = ow.apply(lambda x: x.NAZWA_UKL if x.RODZAJ_CIECIA == "K" else "", axis=1)

    wb = openpyxl.Workbook()
    sheet = wb.active

    sheet["A1"] = "ZLECENIE PRODUKCYJNE: WYDZIAŁ ROZKRÓJ OWAT / KOMPLETACJA OWATY"
    sheet["A3"] = f"NR PARTII: {nr_partii}"
    sheet["A4"] = f"NR ZAMÓWIENIA: {model} {nr_kompletacji}"
    sheet["A5"] = f"MODEL: {model}"
    sheet["H1"] = f"DZIEŃ WYDRUKU      {dt.now().strftime('%Y-%m-%d')}"
    sheet["E4"] = "TYP"
    sheet["F4"] = "MB"
    sheet["G4"] = "R/K"
    sheet["H4"] = "KATER_UKL"

    start_row = 5
    for r in ow.iterrows():

      sheet.cell(row=r[0]+start_row, column=5, value=r[1].TYP_OWATY)
      sheet.cell(row=r[0]+start_row, column=6, value=r[1].ZUZYCIE)
      sheet.cell(row=r[0]+start_row, column=7, value=r[1].RODZAJ_CIECIA)
      sheet.cell(row=r[0]+start_row, column=8, value=r[1].KATER_UKL)


    sheet["A15"] = "BRYŁA"
    sheet["B15"] = "ILOŚĆ PACZEK"
    sheet["C15"] = "CZAS"
    sheet["A16"] = zp.iloc[p].OPIS
    sheet["B16"] = zp.iloc[p].ILE_ZAMOWIONE

    sheet["A18"] = "UWAGI:"
    sheet["H18"] = "PODPIS, DZIEN:"

    thin = Side(border_style="thin", color="000000")
    for row in sheet[f"E4:i{ow.shape[0]+4}"]:
        for cell in row:
          cell.border = Border(bottom=thin)
    
    import os 

    path_ = f'{rozkroj_pianek}{model}/{model} {nr_kompletacji}/OWATY/'
    _file = f"OWATY {model} {nr_kompletacji} {zp.iloc[p].OPIS.replace(model, '').replace('/','_')}.xlsx"
    
    if not os.path.exists(path_):
        os.makedirs(path_)
        print("path:", path_)
    print(_file)

    wb.save(path_ + _file)

def zlecenia_produkcyjne_pianki_kompletacja(model, nr_kompletacji, nr_partii):
  """
  nr_partii -> numer tygodnia / np.: 07/1
  """
  zp = zam_pianki[(zam_pianki.OPIS.str.contains(model))&(zam_pianki.NR_KOMPLETACJI == nr_kompletacji)][["OPIS", "ILE_ZAMOWIONE"]]

  for p in range(zp.shape[0]):

    wb = openpyxl.Workbook()
    sheet = wb.active

    sheet["A1"] = "ZLECENIE PRODUKCYJNE: KOMPLETACJA PIANKI"
    sheet["A3"] = f"NR PARTII: {nr_partii}"
    sheet["A4"] = f"NR ZAMÓWIENIA: {model} {nr_kompletacji}"
    sheet["A5"] = f"MODEL: {model}"
    sheet["H1"] = f"DZIEŃ WYDRUKU      {dt.now().strftime('%Y-%m-%d')}"

    sheet["A15"] = "BRYŁA"
    sheet["B15"] = "ILOŚĆ PACZEK"
    sheet["C15"] = "CZAS"
    sheet["A16"] = zp.iloc[p].OPIS
    sheet["B16"] = zp.iloc[p].ILE_ZAMOWIONE

    sheet["A18"] = "UWAGI:"
    sheet["H18"] = "PODPIS, DZIEN:"

    import os 

    path_ = f'{rozkroj_pianek}{model}/{model} {nr_kompletacji}/KOMPLETACJA/'
    _file = f"KOMPLETACJA {model} {nr_kompletacji} {zp.iloc[p].OPIS.replace(model, '').replace('/','_')}.xlsx"
  
    if not os.path.exists(path_):
        os.makedirs(path_)
        print("path:", path_)

    print(_file)
    wb.save(path_ + _file)




#specyfikacje i zestawienie pianek w modelach do zamowień
def specyfikacja_zam_vita_xlsx(NR_zam:str, raport_vita:pd.DataFrame):
  wb = openpyxl.Workbook()
  sheet = wb.active

  sheet["B1"] = f"SPECYFIKACJA ZAMÓWIENIE DOS/{NR_zam}"
  sheet["H1"] = dt.now().strftime("%Y-%m-%d")
  sheet["B3"] = "Zamawiajacy"
  sheet["B4"] = "OLTA K.K. Zawistowscy sp. j"
  sheet["B5"] = "Ignatki 40/6"
  sheet["B6"] = "16-001 Kleosin"
  sheet["B7"] = "NIP 966 14 08 783"
  sheet["H3"] = "Dostawca"
  sheet["H4"] = "VITA POLYMERS POLAND SP. Z O.O."
  sheet["A8"] = "LP"
  sheet["I8"] = "UWAGI"

  rows = dataframe_to_rows(raport_vita)
  for r_idx, row in enumerate(rows, 1):
      for c_idx, value in enumerate(row, 1):
          sheet.cell(row=r_idx+7, column=c_idx, value=value)

  thin = Side(border_style="thin", color="000000")
  for row in sheet[f"A9:i{raport_vita.shape[0]+9}"]:
      for cell in row:
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

  wb.save(f"SPECYFIKACJA DO ZAMÓWIENIA DOS {' '.join(NR_zam.split('/'))}.xlsx")

def raport_vita(*args):
  """
  Zesatwaienie zamówionych pianek do VITA ilośiowe i z podziałem na bryły
  """
  
  pianki = [x.vita for x in args]
  zam_pianek_vita = pd.concat(pianki).fillna("")[["TYP", "NUMER", "ilosc", "PROFIL", "OZN", "OPIS", "WYMIAR"]]
  zam_pianek_vita.set_index(pd.Index([x for x in range(1,zam_pianek_vita.shape[0]+1)]),inplace=True)

  podsumowanie_zamowienia_vita = pd.concat(pianki).fillna("").sort_values(by="NUMER")
  podsumowanie_zamowienia_vita.set_index(pd.Index([x for x in range(1,podsumowanie_zamowienia_vita.shape[0]+1)]),inplace=True)

  return zam_pianek_vita, podsumowanie_zamowienia_vita[["TYP", "NUMER", "ilosc", "VOL", "PROFIL", "OZN", "OPIS", "WYMIAR"] + [x for x in podsumowanie_zamowienia_vita.columns if "br" in x]]

def raport_dostarczonych_pianek(cls, nr_dos="", drukuj_excel=False):
  """
  Zesatwianie ilosci pianek do modelu z dostawy
 
  cls -> zinicjalizowana instacja klasy zawierająca odpowiedni model

  nr_dos -> numer dostawy z saturn
  """
  zestawienie_pianek_do_bryly = list()

  for br in cls.bryly.keys():
    df = tab[(tab.MODEL == cls.MODEL) & (tab.BRYLA == br)]
    # df["ILOŚĆ"] = (df.ilosc * cls.bryly[br]).astype(int) #ilosc pianek * ilosc kompletów
    df["ILOŚĆ"] = (df.ilosc) #ilosc piabek w modelu wg dokumnetacja
    df["DOSTARCZONO/UWAGI"] = ""
    df = df[['TYP', 'PRZEZ', 'OR', 'OZN', 'PROFIL', 'NUMER', 'WYMIAR', 'TOLERANCJA', 'ILOŚĆ', 'DOSTARCZONO/UWAGI']].fillna("-")
    df["TOLERANCJA"] = df.TOLERANCJA.apply(lambda x: x.replace("+", "").replace("/", "").replace("-", ""))
    df.set_index(pd.Index([x for x in range(1,df.shape[0]+1)]),inplace=True)
    zestawienie_pianek_do_bryly.append([br, df])

  def drukuj_zestawienie_dla_bryly_xlsx(zpdb_n):
    _df = zpdb_n[1]
    # _df = _df[_df.TYP != "G-401"]
    header = f"{cls.MODEL} {zpdb_n[0]} - {cls.bryly[zpdb_n[0]]:.0f}szt"
    # print(header)
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.oddHeader.left.text = header
    sheet.oddHeader.left.size = 18
    sheet.oddHeader.left.font = "Calibry,Bold"
    sheet.oddHeader.right.text = f"Nr dostawy: {nr_dos}"
    sheet.append(["LP"]+list(_df.columns))
    rows = dataframe_to_rows(_df,header=False)
    for r in list(rows)[1:]:
      sheet.append(r)

    thin = Side(border_style="thin", color="000000")
    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    for row in sheet[f"A1:K{_df.shape[0]+1}"]:
      for cell in row:
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    path_ = f'Z:/160. ROZKRÓJ PIANEK/160.50 DOSTAWY PIANEK/{nr_dos.replace("/","_")}/'
    _file = f"{cls.MODEL[:3]} {zpdb_n[0].replace('/','_')}.xlsx"

    import os
    if not os.path.exists(path_):
      os.makedirs(path_)
      print("path:", path_)
    
    print(_file)
    wb.save(path_ + _file)
    # wb.save(f"{cls.MODEL[:3]} {zpdb_n[0]}.xlsx")

  if drukuj_excel:

    for i in zestawienie_pianek_do_bryly:
      drukuj_zestawienie_dla_bryly_xlsx(i)
    return 0

  return zestawienie_pianek_do_bryly

def Wozki_do_dostawy(dostawca:str, nr_dos:str, zam="ZAM1", obj_wozka=5.5, drukuj_do_excel=False):
  """
  dostawca -> np. PIANPOL

  nr_dos -> 24/0347

  zam -> kolumna w zam pianki z ktrej pobieramy nr_dos, defoult: ZAM1
  """

  _tab = tab.copy()
  _tab["vol"] = _tab.DLUG*_tab.SZER*_tab.WYS/1000_000_000
  _tab["VOL"] = _tab.vol*_tab.ilosc

  with engine.begin() as conn:
    wnd = pd.read_sql(text(f"SELECT KOD, OPIS, MODEL, NR_KOMPLETACJI, ILE_ZAMOWIONE, ZNACZNIK_DOSTAWCY, ZAM1, ZAM2, GALANTERIA, SIEDZISKA_HR, LENIWA, UWAGI, nr_SAMOCHODU from ZAM_PIANKI WHERE {zam} = '{nr_dos}'"), conn)

  wnd = wnd.merge(komplety_pianek[["KOD", "BRYLA_GEN", "obj"]], how="left", on="KOD")



  def obj_typ(m,bg,i,g,s,l):

      war_len = (_tab.TYP == "G-401")
      war_shr = (_tab.TYP.str.contains("HR|EE"))

      gal = _tab[(_tab.MODEL == m) & (_tab.BRYLA == bg) & ~war_len & ~war_shr].VOL.sum()*i if g == dostawca[0] else 0
      shr = _tab[(_tab.MODEL == m) & (_tab.BRYLA == bg) & (war_shr)].VOL.sum()*i if s == dostawca[0] else 0
      len = _tab[(_tab.MODEL == m) & (_tab.BRYLA == bg) & (war_len)].VOL.sum()*i if l == dostawca[0] else 0

      # return f"{gal} + {shr} + {len}"
      return gal + shr + len


  wnd["VOL_DOSTAWA"] = wnd.apply(lambda x: obj_typ(x.MODEL, x.BRYLA_GEN, x.ILE_ZAMOWIONE, x.GALANTERIA, x.SIEDZISKA_HR, x.LENIWA), axis=1).round(1)
  wnd["ILE_WOZKOW"] = (wnd.VOL_DOSTAWA / obj_wozka)
  wnd["ILE_WOZKOW"] = wnd["ILE_WOZKOW"].apply(np.ceil).astype(int)
  wnd["OBJ_KOMPLETACJA"] = wnd.ILE_ZAMOWIONE * wnd.obj

  if drukuj_do_excel:
    wnd[["MODEL", "NR_KOMPLETACJI", "OPIS", "ILE_ZAMOWIONE", "VOL_DOSTAWA", "ILE_WOZKOW"]].to_excel(f"WOZKI_{dostawca}_{nr_dos.replace('/', '_')}.xlsx")
  else:
    return wnd
  

def Dodaj_pozycje_do_ZAM_PIANKI(tydzien, zancznik_dostawcy, nr_kompletacji, modele:pd.DataFrame, klasa, zam1=None, zam2=None, nr_partii=None, DODAJ_DO_BAZY = False):

  model = klasa.MODEL
  galanteria = klasa.galanteria
  siedziska_HR = klasa.siedziska_HR
  leniwa = klasa.leniwa



  if DODAJ_DO_BAZY:

    for r in modele.iterrows():
      session.add(ZAM_PIANKI(tydzien, model, r[1].KOD, r[1].OPIS, r[1].DO_ZAMOWIENIA, zancznik_dostawcy, galanteria, siedziska_HR, leniwa, nr_kompletacji, zam1, zam2, "nr_partii: " + nr_partii))

    session.commit()

  else:
    pozycje = list()

    for r in modele.iterrows():
      # tydzien, model, kod, opis, ile_zam, znacznik_dostawcy, galanteria, siedziska_HR, leniwa, nr_kompletacji=None, zam1=None, zam2=None, nr_partii=None
      pozycje.append([tydzien, model, r[1].KOD, r[1].OPIS, r[1].DO_ZAMOWIENIA, zancznik_dostawcy, galanteria, siedziska_HR, leniwa, nr_kompletacji, zam1, zam2, "nr_partii: " + nr_partii])
    
    return pd.DataFrame(pozycje)


def pobierz_zamowienie_z_ZAM_PIANKI(tydzien, _cls):
  _zam = zam_pianki[(zam_pianki.TYDZIEN == tydzien) & (zam_pianki.OPIS.str.contains(_cls.MODEL))].merge(komplety_pianek[["KOD", "BRYLA_GEN"]])[["OPIS", "BRYLA_GEN", "ILE_ZAMOWIONE"]]
  return _cls({i[1].BRYLA_GEN: i[1].ILE_ZAMOWIONE for i in _zam[["BRYLA_GEN", "ILE_ZAMOWIONE"]].iterrows()})